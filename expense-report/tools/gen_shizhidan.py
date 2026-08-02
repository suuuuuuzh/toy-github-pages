#!/usr/bin/env python3
# 按公司「实支单」模板生成报销单 Excel（格式参照 廿一影视 实支单公式版）。
# 用法：python3 tools/gen_shizhidan.py <bakefiles.json> <输出目录>
import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

FONT = "宋体"
THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COMPANY = "廿一影视文化传播（上海）有限公司"
PROJECT = "电影《竹林遗录》"
WIDTHS = {"A": 5.8, "B": 12.3, "C": 12.7, "D": 9.5, "E": 38.0, "F": 9.2, "G": 12.5, "H": 11.3}

# 大写金额公式（与模板一致），{r} 换成合计所在行
UPPER = ('=IF(ROUND(G{r},2)<0,"无效数值",IF(ROUND(G{r},2)=0,"零",IF(ROUND(G{r},2)<1,"",TEXT(INT(ROUND(G{r},2)),"[dbnum2]")&"元")'
         '&IF(INT(ROUND(G{r},2)*10)-INT(ROUND(G{r},2))*10=0,IF(INT(ROUND(G{r},2))*(INT(ROUND(G{r},2)*100)-INT(ROUND(G{r},2)*10)*10)=0,"","零"),'
         'TEXT(INT(ROUND(G{r},2)*10)-INT(ROUND(G{r},2))*10,"[dbnum2]")&"角")&IF((INT(ROUND(G{r},2)*100)-INT(ROUND(G{r},2)*10)*10)=0,"整",'
         'TEXT((INT(ROUND(G{r},2)*100)-INT(ROUND(G{r},2)*10)*10),"[dbnum2]")&"分")))')




def cnupper(amount):
    """人民币中文大写：29703.80 -> 贰万玖仟柒佰零叁元捌角整"""
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["", "拾", "佰", "仟"]
    groups = ["", "万", "亿", "万亿"]
    cents = int(round(float(amount) * 100))
    if cents == 0:
        return "零元整"
    yuan, jiao, fen = cents // 100, (cents // 10) % 10, cents % 10

    def seg4(seg):
        s = ""
        zero_pending = False
        for i in range(3, -1, -1):
            d = (seg // 10 ** i) % 10
            if d:
                if zero_pending:
                    s += "零"
                    zero_pending = False
                s += digits[d] + units[i]
            elif s:
                zero_pending = True
        return s

    segs = []
    y = yuan
    while y > 0:
        segs.insert(0, y % 10000)
        y //= 10000
    res = ""
    for idx, seg in enumerate(segs):
        if seg == 0:
            continue
        g = len(segs) - 1 - idx
        part = seg4(seg)
        if idx > 0 and seg < 1000:
            part = "零" + part
        res += part + groups[g]
    out = (res + "元") if yuan else ""
    if jiao == 0 and fen == 0:
        out += "整"
    else:
        if jiao:
            out += digits[jiao] + "角"
        elif yuan:
            out += "零"
        if fen:
            out += digits[fen] + "分"
        else:
            out += "整"
    return out


def frame(ws, coord):
    """只给（可能是合并区里的）单元格描边，不写值。"""
    ws[coord].border = BOX


def put(ws, coord, value, size=11, bold=False, align="center", box=False, wrap=False, fmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=size, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if box:
        c.border = BOX
    if fmt:
        c.number_format = fmt
    return c


def build(ws, title, rows, loan, report_date, total, loan_note=None):
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E2")
    put(ws, "A1", COMPANY, size=18, bold=True, align="left")
    put(ws, "A4", PROJECT, size=12, bold=True, align="left")
    put(ws, "F4", "实支单编号：", align="left")
    put(ws, "A5", "实支单", size=12, bold=True, align="left")
    put(ws, "F5", "报销日期：", align="left")
    put(ws, "G5", report_date, align="left", fmt="yyyy/m/d")
    put(ws, "A7", "部门：", align="left")
    put(ws, "D7", "姓名：水素", align="left")
    put(ws, "F7", "职务：", align="left")
    ws.merge_cells("G7:H7")
    put(ws, "A8", title, size=11, bold=True, align="left")

    head = 9
    ws.merge_cells(f"D{head}:E{head}")
    for coord, text in (("A", "序号"), ("B", "支付日期"), ("C", "景名/人物"), ("D", "支出项目用途（请填写清楚）"),
                        ("F", "单据数量"), ("G", "金额"), ("H", "预算编号")):
        put(ws, f"{coord}{head}", text, box=True)
    frame(ws, f"E{head}")

    r = head + 1
    for idx, (date, cat, desc, amount, ndocs) in enumerate(rows, 1):
        ws.merge_cells(f"D{r}:E{r}")
        put(ws, f"A{r}", idx, box=True)
        put(ws, f"B{r}", date, box=True, fmt="yyyy/m/d" if isinstance(date, datetime) else None)
        put(ws, f"C{r}", cat, box=True)
        put(ws, f"D{r}", desc, box=True, align="left", wrap=True)
        frame(ws, f"E{r}")
        put(ws, f"F{r}", ndocs, box=True)
        put(ws, f"G{r}", amount, box=True, fmt="#,##0.00")
        put(ws, f"H{r}", "", box=True)
        ws.row_dimensions[r].height = 20
        r += 1

    last = r - 1
    total_row = r + 1
    put(ws, f"A{total_row}", "付款总额：(A) 人民币大写：", align="left")
    # 大写按生成时的合计写成固定文字（用户要求）；右边 G 列数字仍是 SUM 公式
    put(ws, f"D{total_row}", cnupper(total), align="left", bold=True)
    put(ws, f"F{total_row}", "￥：", align="right")
    put(ws, f"G{total_row}", f"=SUM(G{head + 1}:G{last})", bold=True, fmt="#,##0.00", box=True)

    r = total_row + 1
    put(ws, f"A{r}", "付款方式：", align="left")
    put(ws, f"D{r}", "借款金额（如有借款）：（B)", align="left")
    ws.merge_cells(f"G{r}:H{r}")
    put(ws, f"G{r}", loan, fmt="#,##0.00")
    loan_row = r

    r += 1
    put(ws, f"B{r}", "1、现金", align="left")
    put(ws, f"D{r}", "净付金额：(A)-(B)", align="left")
    ws.merge_cells(f"G{r}:H{r}")
    put(ws, f"G{r}", f"=IF(G{total_row}-G{loan_row}>0,G{total_row}-G{loan_row},0)", fmt="#,##0.00")

    r += 1
    put(ws, f"B{r}", "2、支票", align="left")
    put(ws, f"D{r}", "退还借款金额：(B)-(A)", align="left")
    ws.merge_cells(f"G{r}:H{r}")
    put(ws, f"G{r}", f"=IF(G{loan_row}-G{total_row}>0,G{loan_row}-G{total_row},0)", fmt="#,##0.00")

    r += 1
    put(ws, f"B{r}", "3、汇款", align="left")
    put(ws, f"D{r}", "备注：" + (loan_note or ""), align="left")
    ws.merge_cells(f"G{r}:H{r}")

    r += 2
    ws.merge_cells(f"A{r}:B{r}")
    put(ws, f"A{r}", "制片主任/日期：", align="left")
    put(ws, f"E{r}", "部门主管/日期：", align="left")
    put(ws, f"F{r}", "经办人/日期：", align="left")
    r += 1
    put(ws, f"A{r}", "执行制片人/日期：", align="left")
    put(ws, f"E{r}", "制片人/日期：", align="left")
    put(ws, f"F{r}", "总制片人/日期：", align="left")
    r += 1
    put(ws, f"A{r}", "财务审核/日期：", align="left")
    put(ws, f"E{r}", "财务制单/日期：", align="left")
    put(ws, f"F{r}", "领款人/日期：", align="left")


def to_rows(raw):
    out = []
    for iid, date, desc, amount, invamt, cat, files in raw:
        d = None
        if date:
            try:
                d = datetime.strptime(date.strip(), "%Y-%m-%d")
            except ValueError:
                d = date
        out.append((d, cat or "", (desc or cat or "").strip(), amount or 0, len(files)))
    return out


def main():
    bake = json.load(open(sys.argv[1], encoding="utf-8"))
    outdir = sys.argv[2]
    os.makedirs(outdir, exist_ok=True)
    today = datetime(2026, 8, 2)
    LOAN = 52960  # 两次行程合并的借款额（用户提供）
    NOTE = "借款¥52,960按两次行程合并核算，见「汇总」表"

    ka, sa = bake["k"], bake["s"]
    grand = round(ka["total"] + sa["total"], 2)

    wb = Workbook()
    # 汇总 sheet
    ws = wb.active
    ws.title = "汇总"
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:E2")
    put(ws, "A1", COMPANY, size=18, bold=True, align="left")
    put(ws, "A4", PROJECT, size=12, bold=True, align="left")
    put(ws, "A5", "实支单 · 汇总（两次行程合并为一次报销）", size=12, bold=True, align="left")
    put(ws, "F5", "报销日期：", align="left")
    put(ws, "G5", today, align="left", fmt="yyyy/m/d")
    put(ws, "A7", "姓名：水素", align="left")
    r = 9
    for coord, text in (("A", "序号"), ("B", "报销单"), ("D", "笔数"), ("G", "金额")):
        put(ws, f"{coord}{r}", text, box=True)
    for c in ("C", "E", "F", "H"):
        frame(ws, f"{c}{r}")
    rows_spec = [
        ("A-堪景5月", len(ka["rows"])),
        ("B-上海杭州6月", len(sa["rows"])),
    ]
    r += 1
    first = r
    for i, (sheet, cnt) in enumerate(rows_spec, 1):
        put(ws, f"A{r}", i, box=True)
        ws.merge_cells(f"B{r}:C{r}")
        put(ws, f"B{r}", sheet, box=True, align="left")
        frame(ws, f"C{r}")
        put(ws, f"D{r}", cnt, box=True)
        # 引用各 sheet 的合计公式格
        total_row_ref = 9 + cnt + 2  # head=9, 数据从10起，合计在最后一行+2
        put(ws, f"G{r}", f"='{sheet}'!G{total_row_ref}", box=True, fmt="#,##0.00")
        for c in ("E", "F", "H"):
            frame(ws, f"{c}{r}")
        r += 1
    put(ws, f"D{r}", "实际花销合计：(A)", align="left", bold=True)
    put(ws, f"G{r}", f"=SUM(G{first}:G{r-1})", bold=True, fmt="#,##0.00", box=True)
    total_ref = r
    r += 1
    put(ws, f"D{r}", "借款金额：(B)", align="left")
    put(ws, f"G{r}", LOAN, fmt="#,##0.00", box=True)
    loan_ref = r
    r += 1
    put(ws, f"D{r}", "净付金额（公司应补付）：(A)-(B)", align="left")
    put(ws, f"G{r}", f"=IF(G{total_ref}-G{loan_ref}>0,G{total_ref}-G{loan_ref},0)", fmt="#,##0.00", box=True)
    r += 1
    put(ws, f"D{r}", "退还借款金额：(B)-(A)", align="left")
    put(ws, f"G{r}", f"=IF(G{loan_ref}-G{total_ref}>0,G{loan_ref}-G{total_ref},0)", fmt="#,##0.00", box=True)
    r += 2
    put(ws, f"A{r}", "合计大写：", align="left")
    ws.merge_cells(f"B{r}:F{r}")
    put(ws, f"B{r}", cnupper(grand), align="left", bold=True)
    r += 1
    put(ws, f"A{r}", "净付大写：", align="left")
    ws.merge_cells(f"B{r}:F{r}")
    put(ws, f"B{r}", cnupper(max(grand - LOAN, 0)), align="left", bold=True)
    r += 2
    put(ws, f"A{r}", "说明：大写金额按生成时数字写定；如改动明细金额请让我重新生成。", size=9, align="left")

    # A / B 明细 sheet（各自借款 0，备注指向汇总）
    ws_a = wb.create_sheet("A-堪景5月")
    build(ws_a, "堪景（贵州+山东）2026年5月", to_rows(ka["rows"]), 0, today, ka["total"], NOTE)
    ws_b = wb.create_sheet("B-上海杭州6月")
    build(ws_b, "上海+杭州 2026年6月", to_rows(sa["rows"]), 0, today, sa["total"], NOTE)
    # 合并明细 sheet：全部条目 + 合并借款
    ws_m = wb.create_sheet("合并明细")
    allrows = to_rows(ka["rows"]) + to_rows(sa["rows"])
    build(ws_m, "堪景（5月）+ 上海杭州（6月）合并", allrows, LOAN, today, grand, "借款为两次行程合并额")

    path = os.path.join(outdir, "实支单-竹林遗录报销.xlsx")
    wb.save(path)
    print("wrote", path, "| A:", len(ka["rows"]), "B:", len(sa["rows"]), "| grand:", grand,
          "| 净付:", round(max(grand - LOAN, 0), 2))


if __name__ == "__main__":
    main()
